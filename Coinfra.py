import pandas as pd
import unicodedata
import os
import sys
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ==========================================================
# CONFIGURAÇÕES - lidas do arquivo .env (não versionado)
# ==========================================================

load_dotenv()  # carrega as variáveis do arquivo .env para o ambiente


def obter_env(nome, obrigatorio=True, padrao=None):
    """Lê uma variável de ambiente, avisando com clareza se ela estiver faltando."""
    valor = os.getenv(nome, padrao)
    if obrigatorio and not valor:
        sys.exit(
            f"❌ A variável '{nome}' não foi encontrada. Verifique se o arquivo "
            f".env existe e se essa variável está preenchida nele."
        )
    return valor


# --- Arquivos de entrada/saída ---
CAMINHO_PRINCIPAL = obter_env("CAMINHO_PRINCIPAL")
ABA_PRINCIPAL = 0                                  # nome ou índice da aba
COLUNA_MUNICIPIO_PRINCIPAL = "Município"           # nome da coluna de município na principal

CAMINHO_AUXILIAR = obter_env("CAMINHO_AUXILIAR")
ABA_AUXILIAR = 0                                   # nome ou índice da aba
COLUNA_MUNICIPIO_AUXILIAR = "MUNICIPIO"            # nome da coluna de município na auxiliar
COLUNA_CN_AUXILIAR = "CN"                          # coluna que contém o DDD na auxiliar

CAMINHO_SAIDA = obter_env("CAMINHO_SAIDA", padrao="planilha_principal_com_ddd.xlsx")

# --- Filtros a aplicar na planilha principal ---
COLUNA_UF = "UF"                                   # nome da coluna de UF
UF_FILTRO = "BA"                                   # valor de UF desejado

COLUNA_RESPONSAVEL = "Responsável"                 # nome da coluna de responsável
RESPONSAVEIS_FILTRO = ["ICOMON", "O&M ACESSO"]

# --- Configurações do banco de dados MariaDB (lidas do .env) ---
DB_HOST = obter_env("DB_HOST")
DB_PORT = int(obter_env("DB_PORT", padrao="3306"))
DB_USER = obter_env("DB_USER")
DB_SENHA = obter_env("DB_SENHA")
DB_NOME = obter_env("DB_NOME")
DB_TABELA = obter_env("DB_TABELA")   # tabela onde o histórico será gravado

# Colunas que de fato existem na tabela do MariaDB e devem ser enviadas.
# Isso evita erro quando a planilha de origem tem colunas duplicadas
# (o pandas renomeia a segunda ocorrência para "Nome.1", "Nome.2" etc,
# e essas colunas extras não existem na tabela).
# Ajuste esta lista para bater exatamente com as colunas da sua tabela.
COLUNAS_PARA_BANCO = [
    "TA", "Raiz", "Regional", "UF", "Site_Central", "Tipo de Site", "Tipo de Site2",
    "Data Criação", "Alarme", "Responsável", "Nome Site", "Município",
    "Tempo", "Teve Impacto de serviço?", "Evento de impacto Total",
    "Planta Res.", "DDD",
]

# ==========================================================


def normalizar(texto):
    if pd.isna(texto):
        return ""

    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))

    return texto


def filtrar_principal(df):
    for col in (COLUNA_UF, COLUNA_RESPONSAVEL):
        if col not in df.columns:
            raise ValueError(
                f'A coluna "{col}" não foi encontrada na planilha principal. '
                f"Colunas disponíveis: {list(df.columns)}"
            )

    filtro_uf = df[COLUNA_UF].astype(str).str.strip().str.upper() == UF_FILTRO.upper()

    responsavel_norm = df[COLUNA_RESPONSAVEL].astype(str).str.upper()

    valores_norm = [v.upper() for v in RESPONSAVEIS_FILTRO]

    filtro_resp = pd.Series(False, index=df.index)

    for valor in valores_norm:
        filtro_resp |= responsavel_norm.str.contains(valor, na=False)

    df_filtrado = df[filtro_uf & filtro_resp].copy()

    print(
        f"🔎 Filtro aplicado: UF = '{UF_FILTRO}' e Responsável contendo apenas "
        f"{RESPONSAVEIS_FILTRO}. {len(df_filtrado)} de {len(df)} registros mantidos."
    )

    return df_filtrado

def salvar_como_tabela_excel(caminho, nome_tabela="TabelaPrincipal"):
    """
    Abre o arquivo Excel já salvo e converte todo o intervalo de dados em uma
    Tabela do Excel de verdade (com filtro automático, faixa de cores nas
    linhas e nome de tabela) - o mesmo efeito de selecionar os dados no Excel
    e clicar em "Formatar como Tabela".
    """
    wb = load_workbook(caminho)
    ws = wb.active
 
    max_linha = ws.max_row
    max_coluna = ws.max_column
    ultima_coluna = get_column_letter(max_coluna)
    intervalo = f"A1:{ultima_coluna}{max_linha}"
 
    # Garante um nome de tabela válido (sem espaços/caracteres especiais)
    nome_tabela_valido = "".join(c if c.isalnum() or c == "_" else "_" for c in nome_tabela)
    if not nome_tabela_valido or not nome_tabela_valido[0].isalpha():
        nome_tabela_valido = f"T_{nome_tabela_valido}"
 
    tabela = Table(displayName=nome_tabela_valido, ref=intervalo)
    tabela.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tabela)
 
    # Aplica o formato de exibição correto nas colunas TA, Raiz (número) e
    # Data Criação (data e hora), percorrendo o cabeçalho para achar a coluna certa
    colunas_numero = {"TA", "Raiz", "Evento de impacto Total"}
    coluna_data = "Data Criação"
 
    cabecalhos = {celula.value: celula.column for celula in ws[1]}
 
    for nome_coluna in colunas_numero:
        if nome_coluna in cabecalhos:
            letra = get_column_letter(cabecalhos[nome_coluna])
            for celula in ws[letra][1:]:  # pula o cabeçalho
                celula.number_format = "0"
 
    if coluna_data in cabecalhos:
        letra = get_column_letter(cabecalhos[coluna_data])
        for celula in ws[letra][1:]:  # pula o cabeçalho
            celula.number_format = "DD/MM/YYYY"
 
    # Ajusta a largura das colunas automaticamente, já que virou tabela
    for coluna in ws.columns:
        maior_valor = max((len(str(celula.value)) for celula in coluna if celula.value is not None), default=10)
        ws.column_dimensions[coluna[0].column_letter].width = min(maior_valor + 2, 40)
 
    wb.save(caminho)
    print(f"📊 Intervalo de dados formatado como Tabela do Excel ('{nome_tabela_valido}').")
 


def ajustar_tipos_colunas(df):
    df = df.copy()
 
    for coluna in ("TA", "Raiz","Evento de impacto Total" ):
        if coluna in df.columns:
            antes = df[coluna].notna().sum()
            df[coluna] = pd.to_numeric(df[coluna], errors="coerce")
            depois = df[coluna].notna().sum()
            if depois < antes:
                print(
                    f"⚠️  {antes - depois} valor(es) da coluna '{coluna}' não "
                    f"eram numéricos e viraram vazio. Verifique a planilha de origem."
                )
 
    if "Data Criação" in df.columns:
        antes = df["Data Criação"].notna().sum()
        df["Data Criação"] = pd.to_datetime(df["Data Criação"], errors="coerce").dt.normalize()
        depois = df["Data Criação"].notna().sum()
        if depois < antes:
            print(
                f"⚠️  {antes - depois} valor(es) da coluna 'Data Criação' não "
                f"eram datas válidas e viraram vazio. Verifique a planilha de origem."
            )
 
    return df

    


def main():
    df_principal = pd.read_excel(CAMINHO_PRINCIPAL, sheet_name=ABA_PRINCIPAL)
    df_auxiliar = pd.read_excel(CAMINHO_AUXILIAR, sheet_name=ABA_AUXILIAR)

    df_principal = filtrar_principal(df_principal)

    for col, df, nome in [
        (COLUNA_MUNICIPIO_PRINCIPAL, df_principal, "principal"),
        (COLUNA_MUNICIPIO_AUXILIAR, df_auxiliar, "auxiliar"),
        (COLUNA_CN_AUXILIAR, df_auxiliar, "auxiliar"),
    ]:
        if col not in df.columns:
            raise ValueError(
                f'A coluna "{col}" não foi encontrada na planilha {nome}. '
                f"Colunas disponíveis: {list(df.columns)}"
            )

    # Cria uma coluna normalizada só para fazer o "de-para" com segurança
    df_auxiliar["_municipio_norm"] = df_auxiliar[COLUNA_MUNICIPIO_AUXILIAR].apply(normalizar)
    df_principal["_municipio_norm"] = df_principal[COLUNA_MUNICIPIO_PRINCIPAL].apply(normalizar)

    # Monta o dicionário município -> DDD a partir da planilha auxiliar
    # Se houver municípios duplicados na auxiliar, mantém a primeira ocorrência
    mapa_ddd = (
        df_auxiliar.drop_duplicates(subset="_municipio_norm", keep="first")
        .set_index("_municipio_norm")[COLUNA_CN_AUXILIAR]
        .to_dict()
    )

    # Aplica o mapeamento para criar a coluna DDD na planilha principal
    df_principal["DDD"] = df_principal["_municipio_norm"].map(mapa_ddd)

    # Mostra quantos municípios não foram encontrados (para você conferir)
    nao_encontrados = df_principal[df_principal["DDD"].isna()][COLUNA_MUNICIPIO_PRINCIPAL].unique()
    if len(nao_encontrados) > 0:
        print(f"⚠️  {len(nao_encontrados)} município(s) sem DDD correspondente na auxiliar:")
        for m in nao_encontrados:
            print(f"   - {m}")
    else:
        print("✅ Todos os municípios foram encontrados na planilha auxiliar.")

    # Remove a coluna auxiliar de normalização antes de salvar
    df_principal = df_principal.drop(columns=["_municipio_norm"])

    df_principal = ajustar_tipos_colunas(df_principal)

    # Salva o resultado localmente
    df_principal.to_excel(CAMINHO_SAIDA, index=False)
    print(f"\n📄 Arquivo salvo em: {CAMINHO_SAIDA}")

    salvar_como_tabela_excel(CAMINHO_SAIDA, nome_tabela="TabelaTASInfra")


if __name__ == "__main__":
    main()